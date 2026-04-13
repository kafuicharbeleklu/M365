import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df = pd.read_csv('data.csv')
if 'TypeLicence' in df.columns:
    df = df.drop(columns=['TypeLicence'])

# Custom logic: Prioritize "Désactivé" as a specific top-level group
df['GroupCat'] = df['Statut AD'].apply(lambda x: 'COMPTES DESACTIVES' if x == 'Désactivé' else df.loc[df.index == df.index[0], 'EtatActivite'].iloc[0]) 
# Re-adjusting logic for group mapping
def get_group(row):
    if row['Statut AD'] == 'Désactivé':
        return 'COMPTES DESACTIVES'
    return row['EtatActivite']

df['GroupCat'] = df.apply(get_group, axis=1)
df = df.sort_values(by=['GroupCat', 'EtatActivite'])

wb = Workbook()
ws = wb.active
ws.title = 'Reporting'

hex_color = 'C09F29'
header_fill = PatternFill(start_color=hex_color, end_color='C09F29', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
section_fill = PatternFill(start_color='444444', end_color='444444', fill_type='solid')
section_font = Font(color='FFFFFF', bold=True, size=12)
total_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
total_font = Font(bold=True)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

current_row = 1
for cat, group in df.groupby('GroupCat', sort=False):
    current_row += 2
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(df.columns)-1)
    cell = ws.cell(row=current_row, column=1, value=f'SESSION : {cat.upper()}')
    cell.fill = section_fill
    cell.font = section_font
    current_row += 1
    
    # Header
    for c, col in enumerate(df.columns[:-1], 1):
        cell = ws.cell(row=current_row, column=c, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    current_row += 1
    
    # Data
    for _, row in group.drop(columns=['GroupCat']).iterrows():
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=c, value=val)
            cell.border = border
        current_row += 1
    
    # Total
    cell_label = ws.cell(row=current_row, column=len(df.columns)-2, value='TOTAL SECTION')
    cell_label.fill = total_fill; cell_label.font = total_font; cell_label.border = border; cell_label.alignment = Alignment(horizontal='right')
    val_cell = ws.cell(row=current_row, column=len(df.columns)-1, value=len(group))
    val_cell.fill = total_fill; val_cell.font = total_font; val_cell.border = border; val_cell.alignment = Alignment(horizontal='center')

for i in range(1, len(df.columns)):
    col_letter = get_column_letter(i)
    max_length = 0
    for row in ws.iter_rows(min_col=i, max_col=i):
        for cell in row:
            if cell.value: max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 4

wb.save('Rapport_Activite_Utilisateurs.xlsx')
print('Fichier généré : Rapport_Activite_Utilisateurs.xlsx')
